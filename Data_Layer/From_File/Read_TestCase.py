#-*- coding:utf-8 -*-#
#-------------------------------------------------------------------------
#ProjectName:       Python2020
#FileName:          Read_TestCase.py
#Author:            Ben
#Date:              2020/8/1 16:15
#Description:实现测试用例读取操作并完成代码封装，要将excel中的每行每列进行提取出来
#--------------------------------------------------------------------------
import openpyxl
from Data_Layer.EXCEL_CONSTANT import *
from Data_Layer.From_File.Read_Yaml import ReadYaml
from Common_Layer.PATH_CONSTANTS import EXCEL_PATH,EXPECT_PATH,PARAMS_PATH
class  ReadTestCase(object):
    def __init__(self):
        #获取操作excel的对象,需要传入execl文件所在路径
        self.read_excel=openpyxl.load_workbook(EXCEL_PATH)
        #获取到所有的sheet
        self.get_sheets=self.read_excel.sheetnames

    def get_cell_value(self,column,sheet_name,row):
        cell_coord=column+str(row)
        for key,value in self.get_merge_cells(sheet_name).items():
            if cell_coord in value:
                return self.read_excel[sheet_name][key].value  #如果传入的单元格形式是在我合并单元格的值中的话，那么我直接返回键的值即可
        return self.read_excel[sheet_name][column+str(row)].value
    #获取到请求的url、请求的参数、请求的方法
    #分析发现获取一个单元格要求由行与列所构成，但是如果两个都变动那么维护或者代码实现过于复杂，所以我们就
    #列固定，行不固定，列就会设置为常量
    def get_url(self,sheet_name,row):
        return self.get_cell_value(CASE_URL,sheet_name,row)
    #获取到excel中的方法
    def get_method(self,sheet_name,row):
        return self.get_cell_value(CASE_METHOD,sheet_name,row)
    #获取到excel中的参数
    def get_params(self,sheet_name,row):
        return self.get_cell_value(CASE_PARAMS,sheet_name,row).split(":")
    #获取excel中的最大行
    def get_max_row(self,sheet_name):
        return self.read_excel[sheet_name].max_row  #直接调用sheet对象的max_row属性
    #声明一个方法完成合并单元格的处理
    def get_merge_cells(self,sheet_name):
        list_cell = []
        dict_cell = {}
        for cellrange in self.read_excel[sheet_name].merged_cells.ranges:  # 获取到excel中的所有合并单元格对象
            # 想要的格式：{"C2":["C2","C3","C4"],"C5":["C5","C6","C7"]}
            get_cell=cellrange.coord.split(":")
            key = get_cell[0]  # 合并单元格的coord的形式就是  开始的单元格形式：结束的单元格形式
            if get_cell[0][0]==get_cell[1][0]:  #表示的判定截取出来的两个元素的首字符是否相同
                # 获取每个cellrange的行
                rows = len([value for value in cellrange.rows])
                start_row = int(key[1:])  # 定义一个开始行
                for row in range(start_row, start_row + rows):
                    list_cell.append(key[0] + str(row))
                dict_cell.update({key: list_cell})
                # 每个cellrange存储一个list，那么没存储完之后要将list清空
                list_cell = []
            else:
                columns = len([value for value in cellrange.cols])
                for col in range(ord(key[0]),ord(key[0])+columns):
                    list_cell.append(chr(col) + str(get_cell[0][1:]))
                dict_cell.update({key: list_cell})
                list_cell=[]
        return dict_cell

    #获取是否mouck的值
    def get_ismock(self,sheet_name,row):
        return self.get_cell_value(CASE_MOCK,sheet_name,row)


    #获取是否通过的单元格对象
    def get_ifpass(self,sheet_name,row):
        return self.read_excel[sheet_name][CASE_PASS + str(row)]


    #获取预期结果的值
    def get_expect(self, sheet_name, row):
        return self.get_cell_value(CASE_EXPECT, sheet_name, row).split(":")
    #获取实际结果单元格
    def get_actual(self,sheet_name,row):
        #在excel中的所有赋值的可能性，三种常用的赋值形式
        #1.通过sheet获取具体的cell单元格对象进行赋值操作
        # self.read_excel[sheet_name][CASE_ACTUAL + str(row)]="test"
        # #3.通过cell对象调用value属性进行赋值，在celle对象调用value时拥有另个实现功能，一个是直接获取当前单元格的值，一个是覆盖之前单元格的值
        # self.read_excel[sheet_name][CASE_ACTUAL + str(row)].value="第九列第二行单元格"
        # # print("方法中的",get_value)
        # # self.read_excel[sheet_name][CASE_ACTUAL+str(row)]=value
        # # self.read_excel.save(EXCEL_PATH)
        return self.read_excel[sheet_name][CASE_ACTUAL + str(row)]


    #获取测试用例名
    def get_case_name(self,sheet_name,row):
        return self.get_cell_value(CASE_NAME, sheet_name, row)


    #获取测试是否存在依赖
    def get_depend(self,sheet_name,row):
        return self.get_cell_value(CASE_DEPEND, sheet_name, row)

    #获取依赖cookie的关联字段
    def get_cookie(self, sheet_name, row):
        return self.get_cell_value(CASE_COOKIE, sheet_name, row)
    #获取用例ID
    def get_case_id(self,sheet_name,row):
        return  self.get_cell_value(CASE_ID, sheet_name, row)

    #获取响应依赖的用例ID
    def get_case_response(self,sheet_name,row):
        return  self.get_cell_value(CASE_RESPONSE, sheet_name, row)
    #获取响应依赖的字段
    def get_case_field(self,sheet_name,row):
        return  self.get_cell_value(CASE_FIELD, sheet_name, row)

    #获取请求的字段
    def  get_request(self,sheet_name,row):
        return self.get_cell_value(CASE_REQUST, sheet_name, row)

    def get_session(self, sheet_name, row):
        return self.get_cell_value(CASE_SESSION, sheet_name, row)


if __name__ == '__main__':
    read=ReadTestCase()   #{"C2":["C2","C3","C4"],"C5":["C5","C6","C7"],"G3":["G3","H3"]}
    for cellrange in read.read_excel["Sheet2"].merged_cells.ranges:
        print(cellrange)
    print(read.get_merge_cells("Sheet2"))

# read.get_actual("Sheet1",2)   #表示的是对实际结果的第二行的单元格进行赋值test
    # #第二种赋值的形式,通过sheet调用cell对象传入参数完成赋值，其row和column都是int类型参数
    # read.read_excel["Sheet1"].cell(row=3,column=9,value="第九列第三行单元格")
    # #在main方法中调用value属性进行赋值操作
    # get_cell=read.get_actual("Sheet1",4)  #第四行第九列的单元格的值
    # get_cell.value="第九列第四行单元格值"
    # get_cell_1 = read.get_actual("Sheet1", 5)  # 第五行第九列的单元格的值
    # get_cell_1="第九列第五行单元格的值"   #返回的不是cell对象吗？不是可以直接赋值吗？  只是将我们获取到的cell对象进行转换成了字符串对象，
    # #如果想要通过cell对象直接赋值的话，则需要一步完成，才能够写入excel中。
    # print(type(get_cell_1))
    # #此时get_cell_1是一个cell对象，如果直接复制相当于将字符串对象覆盖了之前的cell对象，压根没有实现cell对象赋值操作
    # read.read_excel.save(EXCEL_PATH)
    #[{a:b},{d:e}]
    #格式固定会用:格式声明数据映射的格式,参数化两个节点参数，一个父节点一个子节点
    # for i in range(2,read.get_max_row("Sheet1")+1):
    #     #print(read.get_method("Sheet1",i))
    #     read.get_actual("Sheet1",i,1111)
    #     get_actul=read.get_actual("Sheet1",i,1111)
    #     print("main方法中：",get_actul)
    #     read.read_excel.save(EXCEL_PATH)




